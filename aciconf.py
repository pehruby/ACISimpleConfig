import yaml
import json
import os
import argparse
import getpass

from openpyxl import load_workbook
from gssacipkg import ACIobj, ACIconfig


def read_config_file(config_file):
    """Read an YAML config file
    :param config_file: [description]
    :type config_file: [type]
    """

    if os.path.isfile(config_file):
        extension = os.path.splitext(config_file)[1]
        try:
            with open(config_file) as data_file:
                if extension in ['.yml', '.yaml']:
                    config_json = yaml.safe_load(data_file.read())
                elif extension in ['.json']:
                    config_json = json.safe_load(data_file.read())
        except IOError:
            print("Unable to read the file", config_file)
            exit(1)
    else:
        print("Cannot find the file", config_file)
        exit(1)

    return config_json


def list2dict(listvar):
    """Transform list to dict
    :param listvar: [description]
    :type listvar: [type]
    :return: [description]
    :rtype: [type]
    """
    resultdict = {}
    for item in listvar:
        key = list(item.keys())[0]
        resultdict[key] = item[key]
    return resultdict


def process_excel_cmd(cmd_line):
    """Process cmd line parameters related to Excel file.
    filename;table;line
    line is line number where column names are defined
    
    :param cmd_line: [description]
    :type cmd_line: [type]
    """
    param_dict = {}
    cmd_line_list = cmd_line.split(';')
    if len(cmd_line_list) != 3:
        print("Bad number of Ecel parameters")
        exit(1)
    for listentry in cmd_line_list:
        if listentry == "":
            print("Excel parameters - wrong format")
            exit(1)
    param_dict["filename"] = cmd_line_list[0]
    param_dict["tab"] = cmd_line_list[1]
    param_dict["paramline"] = cmd_line_list[2]

    return(param_dict)


def get_execel_table(paramdict):
    """Get Excel table based on cmd line parameters
    
    :param paramdict: [description]
    :type paramdict: [type]
    """

    wb = load_workbook(paramdict["filename"])
    sheet = wb[paramdict["tab"]]
    # titles = sheet[1]
    table = readxlstable(sheet, paramdict["paramline"])
    return(table)


def readxlstable(sheet, colnamerow):
    """Read rows from a excel table.
    Result is list of dict, where each list entry contains one 
    Excel line, dict keys are column names
    
    :param sheet: [description]
    :type sheet: [type]
    :param colnamerow: [description]
    :type colnamerow: [type]
    """
    table = []
    title = []
    titles_raw = sheet[colnamerow]
    for item in titles_raw:
        title.append(item.value)
    nr_columns = sheet.max_column

    for line in sheet.iter_rows(min_row=int(colnamerow)+1, values_only=True):
        tableitem = {}
        for item in range(nr_columns):
            # if item value in Excel sheet is empty (None), replace is with ''
            tableitem[title[item]] = line[item] if line[item] is not None else ''

        # print(line)
        table.append(tableitem)
    return(table)


def main():

    username = ""
    pswd = ""
    aci_ip = ""
    config_file = ""
    aci_api_cfg_file = ""
    debug = False
    is_excel = False    # config parameters are in excel file
    excel_cmd_line = ""
    excel_param_dict = {}

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-d", "--debug", help="debug", action="store_true", required=False
    )
    parser.add_argument("-i", "--ip", help="ACI IP address or hostname", required=True)
    parser.add_argument("-u", "--username", help="ACI username", required=True)
    parser.add_argument("-p", "--password", help="ACI password", required=False)
    parser.add_argument("-c", "--config", help="configuration file", required=True)
    parser.add_argument(
        "-a", "--aciapidesc", help="ACI API descriprion configuration", required=True
    )
    parser.add_argument(
        "-D", "--delete", help="delete configuration", action='store_true'
    )
    parser.add_argument(
        "-x", "--excel", help="Excel file parameters", required=False
    )
    parser.add_argument(
        "-t", "--test", help="Parses config file only", action="store_true",
        required=False
    )

    args = parser.parse_args()
    aci_ip = args.ip
    debug = args.debug  # noqa
    username = args.username
    pswd = args.password
    config_file = args.config
    aci_api_cfg_file = args.aciapidesc
    excel_cmd_line = args.excel
    parse_config_only = args.test

    if pswd == None:    # noqa
        pswd = getpass.getpass("Password:")

    # acicfgcode = read_config_file(config_file)  # ACI IaC conf from a yaml file
    aciapilist = read_config_file(
        aci_api_cfg_file
    )  # list of API config URLs for specific ACI items

    aciapidict = list2dict(aciapilist['aciapidesc'])
    if excel_cmd_line:
        # parameters for j2 template are in Excel file
        is_excel = True
        # cmd line excel parameters are in form filename|sheet|line
        # line in line number where column names are defined
        # transfor parameters into a dict
        excel_param_dict = process_excel_cmd(excel_cmd_line)
        # create list of dict, each entry of list contains dict with one line
        # of excel table, keys of the dict are column names
        exceltable = get_execel_table(excel_param_dict)
        # process excel table line by line
        for tableline in exceltable:
            # create structure which is normaly in yaml config file
            cfgstruct = {
                'vars': tableline,
                'aci_cfgfiles': [config_file]   # j2 template file is expected here
            }
            # cfgstruct and not yaml file name is the first parameter if it is_excel
            MyACIconfig = ACIconfig(cfgstruct, aciapidict, is_excel)
            if not parse_config_only:   # not a test (config parsing), deploy config
                MyACI = ACIobj(username, pswd, aci_ip, MyACIconfig, aciapilist)
                MyACI.deploycfg()
    else:
        # standard yaml file configuration
        MyACIconfig = ACIconfig(config_file, aciapidict)
        if not parse_config_only:   # not a test (config parsing), deploy config
            MyACI = ACIobj(username, pswd, aci_ip, MyACIconfig, aciapilist)
            if args.delete:
                MyACI.deletecfg()
            else:
                MyACI.deploycfg()
    print("End")


if __name__ == "__main__":

    main()
